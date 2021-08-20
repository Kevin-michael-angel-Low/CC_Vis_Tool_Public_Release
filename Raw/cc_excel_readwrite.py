import openpyxl

# filename = "CC_Abnormalities_List.xlsx"

# wb = openpyxl.load_workbook(filename)

# Function used to take path and abnormality and append to excel sheet
def add_row(cc_path_for_excel, abnormality_for_excel, image_num_for_excel, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    col_lenA = len(ws["A"])
    col_lenB = len(ws["B"])
    col_lenC = len(ws["C"])
    inSheet = False

    for row in ws.iter_rows(min_row = 1, max_row = col_lenA, max_col = 1):
        for cell in row:
            if cell.value == image_num_for_excel:
                print("Changing existing")
                inSheet = True
                ws['A' + str(cell.row)] = image_num_for_excel
                ws['B' + str(cell.row)] = abnormality_for_excel
                ws['C' + str(cell.row)] = cc_path_for_excel

    if (ws['A' + str(col_lenA)] != image_num_for_excel) and (inSheet ==  False):
        print("Adding new")
        ws['A' + str(col_lenA + 1)] = image_num_for_excel
        ws['B' + str(col_lenB + 1)] = abnormality_for_excel
        ws['C' + str(col_lenC + 1)] = cc_path_for_excel

    wb.save(excel_file)

def view_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    view_table = ''
    excels_max_row = ws.max_row
    if excels_max_row > 15:
        for row in ws.iter_rows(min_col = 1, max_col = 2, min_row = excels_max_row - 14):
            for cell in row:
                view_table += (str(cell.value) + "\n")
        return view_table
    else:
        for row in ws.iter_rows(min_col = 1, max_col = 2, max_row = excels_max_row):
            for cell in row:
                view_table += (str(cell.value) + "\n")
        return view_table

    