import pickle
import xlsxwriter
import openpyxl

# with open ('cc_40k.pkl', 'rb') as filehandle:
#     cc_40k = pickle.load(filehandle)

# with open("cc_files.pkl", 'rb') as filehandle:
#     cc_10k = pickle.load(filehandle)

# print(len(cc_10k))
# print(len(cc_40k))

# cc_30k = [i for i in cc_40k if i not in cc_10k]

# print(len(cc_30k))

# with open ('cc_30k.pkl', 'wb') as filehandle:
#     pickle.dump(cc_30k, filehandle)

with open('cc_40k.pkl', 'rb') as filehandle:
    cc_files = pickle.load(filehandle)
print(cc_files[0])
# print("new 30k windows:"+ str(len(cc_files)))

# wb = openpyxl.Workbook()
# ws = wb.active

# for column in range (0, len(cc_files)):
#     ws['A' + str(column)] = cc_files[column]

# for column in len(cc_files):
#     ws.append(cc_files[column])

# wb.save("new_40k_full.xlsx")


# Pickle file to excel------
workbook  = xlsxwriter.Workbook('cc_40k_full.xlsx')
worksheet = workbook.add_worksheet()
worksheet.write_column('A1', cc_files)

workbook.close()
#-----------------

# cc_cut = [cut[8:] for cut in cc_files]

# added = ["//smb-ifs.ini.usc.edu" + add for add in cc_cut]
# print(added[0])

# with open("cc_30k_win.pkl", "wb") as filehandle:
#     pickle.dump(added, filehandle)


# Working path on cc_final
# Note: To run, first conda activate CCVis
# \\smb-ifs.ini.usc.edu\faculty/njahansh/nerds/alyssa/corpus_callosum/UKBB/bfcorr_results/6026120_20216_2_0/segmentation/6026120_20216_2_0/6026120_20216_2_0_std_BF_adj_flirt.nii.gz

# with open ('cc_40k.pkl', 'wb') as filehandle:
#     pickle.dump(cc_files, filehandle)
# print(cc_files)